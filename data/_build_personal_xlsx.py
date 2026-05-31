"""Build data/personal-touches.xlsx from data/personal.json."""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(ROOT, "personal.json")
OUT = os.path.join(ROOT, "personal-touches.xlsx")

with open(SRC, "r", encoding="utf-8") as f:
    data = json.load(f)

wb = Workbook()
ws = wb.active
ws.title = "Personal Touches"

headers = ["ID", "Type", "Category", "Size", "Title", "Description", "Image", "Tags"]
ws.append(headers)

header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
header_fill = PatternFill("solid", fgColor="FF29404D")
center = Alignment(horizontal="center", vertical="center")
wrap = Alignment(vertical="top", wrap_text=True)
thin = Side(border_style="thin", color="FFCCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for col, _ in enumerate(headers, start=1):
    c = ws.cell(row=1, column=col)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = border

for item in data:
    ws.append([
        item.get("id", ""),
        item.get("type", ""),
        item.get("category", ""),
        item.get("size", ""),
        item.get("title", ""),
        item.get("description", ""),
        item.get("image", ""),
        item.get("tags", ""),
    ])

# Style data rows
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
    for cell in row:
        cell.alignment = wrap
        cell.border = border

widths = [28, 12, 18, 10, 36, 70, 36, 30]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[chr(64 + i)].width = w

ws.row_dimensions[1].height = 22
for r in range(2, ws.max_row + 1):
    ws.row_dimensions[r].height = 70

ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

wb.save(OUT)
print(f"Wrote {OUT} ({len(data)} rows)")
